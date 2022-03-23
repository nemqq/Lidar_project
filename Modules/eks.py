# Library for saving data from lidar to excel
# Author: Kamil Sikora
# Date: 01.01.2022
# <kamil.sikora@studnet.po.edu.pl>


import openpyxl as xl
from openpyxl.chart import LineChart, Reference


def fnc2(table, m_name_p):
    """
    function writes data from list to txt file


    :param table: list
    :param m_name_p: name of .txt file (name must contains .txt extension)
    :return:
    """
    try:
        with open(m_name_p, "r+") as f:
            text = f.readlines()
            counter = (len(text)) + 1
            f.write("Data nr ... : " + str(counter) + "  " + str(table) + "\n")

    except FileNotFoundError:
        test = open(m_name_p, "w")
        test.close()
        print("file with name " + m_name_p + "has been created")


class Excel:
    """
    Class with functions for entering data into Excel
    """

    @staticmethod
    def create(name):
        """
        Create a file with .xlsx
        :param name: file with .xlsx
        :return:
        """

        try:
            wb = xl.load_workbook(name)
            wb.save(name)
            print('File was not created because it does not exist!')

        except:

            wb = xl.Workbook()
            ws = wb.active
            ws.title = 'Data_1'
            wb.save(name)
            print('THE FILE HAS BEEN CREATED SUCCESSFULLY ')

    def save_data_row(self, name, tabe_1, u_row, u_col, case=True):
        """
        Writing data to the sheet

        :param name: file name with extension
        :param tabe_1: data from the board
        :param u_row: first row
        :param u_col: first column
        :param case: True - save perpendicularly, False - save horizontally
        :return:
        """

        wb = xl.load_workbook(name)
        ws = wb.active

        if case is True:
            if u_row == 0:
                t = 1
            else:
                t = u_row
            for row in range(len(tabe_1)):
                ws.cell(row + t, u_col, tabe_1[row])
        if case is False:
            if u_col == 0:
                t = 1
            else:
                t = u_col
            for col in range(len(tabe_1)):
                ws.cell(u_row, col+t, tabe_1[col])

        wb.save(name)

    def save_data_apend(self, name, tabe_1):
        """
        apend

        :param name:  file name
        :param tabe_1:  data from board
        :return:
        """
        wb = xl.load_workbook(name)
        ws = wb.active
        ws.append(tabe_1)
        wb.save(name)

    def create_graph(self, name):
        """
        still in progress fun

        :param name:
        :return:
        """
        wb = xl.load_workbook(name)
        ws = wb.active

        values = Reference(ws, min_col=1, min_row=1, max_col=1, max_row=19)
        chart = LineChart()
        chart.title = 'Dane nr 1'
        chart.style = 1
        # chart.x_axis.title='nr probe'
        # chart.y_axis.title='dist cm '
        chart.add_data(values)
        ws.add_chart(chart, "E15")

        wb.save(name)
